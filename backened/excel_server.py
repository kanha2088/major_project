from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
import os

app = Flask(__name__)
CORS(app)

@app.route("/get-user-data", methods=["POST"])
def get_user_data():
    data = request.get_json()
    user_id = str(data.get("id"))
    request_type = data.get("type")  # "scanner" or "table"

    try: 
        excel_path = r"C:\Users\sneha padhy\OneDrive\Desktop\users2.xlsx"
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]
        id_index = headers.index('Id')
        rice_index = headers.index("Rice Alloted(kgs)")
        months = headers[5:]  # assuming 5th column onwards are months

        user_data = None
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[id_index]) == user_id:
                user_data = row
                break
        print(user_data)
        if not user_data:
            return jsonify({"error": "User not found"}), 404

        total_rice = float(user_data[rice_index])

        if request_type == "scanner":
            # Return minimal data for Scanner
            return jsonify({
                "totalRice": total_rice,
                "year": user_data[3]  # assuming year is at column index 3
            })

        elif request_type == "table":
            # Return full month-wise dispensed data
            month_wise_dispensed = {}
            for i, month in enumerate(months):
                if month is None:
                    continue
                try:
                    value = float(user_data[i + 5]) if user_data[i + 5] is not None else 0
                except (ValueError, IndexError):
                    value = 0
                month_wise_dispensed[month] = value
            print("monthsdata",month_wise_dispensed)
            return jsonify({
                "totalRice": total_rice,
                "year": user_data[3],
                "dispensed": month_wise_dispensed,
                "months": months
            })

        else:
            return jsonify({"error": "Invalid request type"}), 400

    except Exception as e:
        return jsonify({"error": "Server error", "details": str(e)}), 500

if __name__ == "__main__":
    app.run(port=5001, debug=True, use_reloader=False)
